from datetime import date
from decimal import Decimal, InvalidOperation
from functools import wraps

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from accounts.forms import StaffCreateForm
from accounts.models import User
from core.models import RateCard, RestrictedItem, ContactEnquiry
from core.terms import TERMS_AND_CONDITIONS
from tracking.models import Parcel, TrackingEvent
from .models import (AuditLog, Box, BoxCharge, BoxItem, Branch, CUSTOMS_TAX,
                     Expense, Invoice, InvoiceItem, log)
import re
from io import BytesIO

# ---------- permissions ----------

def panel_required(view):
    @wraps(view)
    @login_required
    def wrapper(request, *a, **kw):
        if not request.user.is_panel_user:
            return HttpResponseForbidden('Staff panel access only.')
        return view(request, *a, **kw)
    return wrapper


def owner_required(view):
    @wraps(view)
    @login_required
    def wrapper(request, *a, **kw):
        if not request.user.is_owner:
            return HttpResponseForbidden('Owner access only.')
        return view(request, *a, **kw)
    return wrapper


def branch_scope(qs, user):
    """Owner sees all branches; everyone else only their own."""
    return qs if user.is_owner else qs.filter(branch=user.branch)


# ---------- dashboard ----------

@panel_required
def dashboard(request):
    today = timezone.localdate()
    parcels = branch_scope(Parcel.objects.filter(is_draft=False), request.user)
    invoices = branch_scope(Invoice.objects.all(), request.user)
    drafts = branch_scope(Parcel.objects.filter(is_draft=True), request.user)
    ctx = {
        'today_parcels': parcels.filter(created_at__date=today).count(),
        'total_parcels': parcels.count(),
        'in_transit': parcels.filter(status=Parcel.Status.IN_TRANSIT).count(),
        'today_income': invoices.filter(created_at__date=today).aggregate(s=Sum('grand_total'))['s'] or 0,
        'recent': parcels.select_related('branch')[:8],
        'draft_count': drafts.count(),
        'enquiries': ContactEnquiry.objects.filter(is_resolved=False)[:5] if request.user.is_owner else None,
    }
    return render(request, 'panel/dashboard.html', ctx)


# ---------- new shipment (the heart of the SaaS) ----------

def _parse_customs(post):
    """Customs tax: defaults to the standard NPR 1,800 but the field is
    editable, so a special customer can be given a different amount."""
    raw = (post.get('customs_tax') or '').strip()
    if raw == '':
        return CUSTOMS_TAX
    value = Decimal(raw)
    if value < 0:
        raise ValueError('Customs tax cannot be negative.')
    return value.quantize(Decimal('0.01'))


def _parse_boxes(post, lenient=False):
    """Read box/item/charge rows from the form.
    lenient=True (drafts): empty weights/rates become 0 instead of errors."""
    box_indices = sorted({
        int(m.group(1))
        for k in post.keys()
        for m in [re.match(r'box\[(\d+)\]\[weight\]', k)] if m
    })
    boxes_data = []
    for i in box_indices:
        w_raw = (post.get(f'box[{i}][weight]') or '').strip()
        r_raw = (post.get(f'box[{i}][rate]') or '').strip()
        if lenient:
            weight = Decimal(w_raw) if w_raw else Decimal('0')
            rate = Decimal(r_raw) if r_raw else Decimal('0')
        else:
            weight = Decimal(w_raw)
            rate = Decimal(r_raw)

        item_names = post.getlist(f'box[{i}][item_name][]')
        item_qtys = post.getlist(f'box[{i}][item_qty][]')
        items = [(n.strip(), int(q or 1)) for n, q in zip(item_names, item_qtys) if n.strip()]

        charge_names = post.getlist(f'box[{i}][charge_name][]')
        charge_weights = post.getlist(f'box[{i}][charge_weight][]')
        charge_prices = post.getlist(f'box[{i}][charge_price][]')
        charges = [
            (cn.strip(), Decimal(cw) if cw.strip() else None, Decimal(cp))
            for cn, cw, cp in zip(charge_names, charge_weights, charge_prices)
            if cp.strip()
        ]
        boxes_data.append({'weight': weight, 'rate': rate, 'items': items, 'charges': charges})
    return boxes_data


def _save_boxes(parcel, boxes_data):
    parcel.boxes.all().delete()
    for b in boxes_data:
        box = Box.objects.create(parcel=parcel, weight=b['weight'], rate_per_kg=b['rate'])
        BoxItem.objects.bulk_create([BoxItem(box=box, name=n, quantity=q) for n, q in b['items']])
        BoxCharge.objects.bulk_create([BoxCharge(box=box, name=cn, weight=cw, price=cp)
                                       for cn, cw, cp in b['charges']])


def _fill_parcel_fields(parcel, post):
    parcel.sender_name = post.get('sender_name', '').strip()
    parcel.sender_phone = post.get('sender_phone', '').strip()
    parcel.sender_address = post.get('sender_address', '').strip()
    parcel.sender_email = post.get('sender_email', '').strip()
    parcel.receiver_name = post.get('receiver_name', '').strip()
    parcel.receiver_phone = post.get('receiver_phone', '').strip()
    parcel.receiver_address = post.get('receiver_address', '').strip()
    parcel.destination_country = post.get('destination_country', '').strip()


@panel_required
def shipment_create(request, pk=None):
    """Create a shipment OR continue editing a saved draft (pk given).

    Two save buttons:
      action=draft -> saves everything as a draft (no invoice, no tracking,
                      hidden from the public site) — can be edited freely.
      action=final -> full validation, then parcel + invoice + boxes + QR
                      are created together and pricing is LOCKED forever.
    """
    rates = RateCard.objects.filter(is_active=True)
    draft = None
    if pk is not None:
        draft = get_object_or_404(
            branch_scope(Parcel.objects.filter(is_draft=True), request.user), pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action', 'final')
        try:
            customs = _parse_customs(request.POST)

            if action == 'draft':
                boxes_data = _parse_boxes(request.POST, lenient=True)
                if not request.POST.get('sender_name', '').strip() and \
                   not request.POST.get('sender_phone', '').strip():
                    raise ValueError('Enter at least the sender name or phone to save a draft.')
                with transaction.atomic():
                    parcel = draft or Parcel(
                        tracking_id=Parcel.next_tracking_id(),
                        branch=request.user.branch or Branch.objects.filter(is_active=True).first(),
                        created_by=request.user, is_draft=True)
                    _fill_parcel_fields(parcel, request.POST)
                    parcel.is_draft = True
                    parcel.draft_customs_tax = customs
                    parcel.save()
                    _save_boxes(parcel, boxes_data)
                    log(request.user, f'Draft saved: {parcel.tracking_id} ({parcel.sender_name or parcel.sender_phone})')
                messages.success(request, f'Draft {parcel.tracking_id} saved — finish it any time from the Drafts page.')
                return redirect('drafts')

            # ----- final save -----
            country = request.POST['destination_country']
            rates.get(country=country)  # validates the destination has a rate card
            boxes_data = _parse_boxes(request.POST, lenient=False)
            if not boxes_data:
                raise ValueError('Add at least one box.')
            if not any(b['items'] for b in boxes_data):
                raise ValueError('Add at least one item.')

            total_weight = sum(b['weight'] for b in boxes_data)
            weight_charge = sum((b['weight'] * b['rate']).quantize(Decimal('0.01')) for b in boxes_data)
            additional_charges_total = sum(
                (c[2] for b in boxes_data for c in b['charges']), Decimal('0.00'))
            grand_total = weight_charge + additional_charges_total + customs

            branch = request.user.branch or Branch.objects.filter(is_active=True).first()

            with transaction.atomic():  # all-or-nothing
                parcel = draft or Parcel(
                    tracking_id=Parcel.next_tracking_id(), branch=branch,
                    created_by=request.user)
                _fill_parcel_fields(parcel, request.POST)
                parcel.is_draft = False
                parcel.draft_customs_tax = None
                if not parcel.qr_code:
                    parcel.generate_qr()
                parcel.save()
                _save_boxes(parcel, boxes_data)

                seq_year = timezone.now().year
                last = Invoice.objects.filter(invoice_no__startswith=f'INV-{seq_year}-').order_by('-id').first()
                seq = int(last.invoice_no.split('-')[-1]) + 1 if last else 1
                invoice = Invoice.objects.create(
                    parcel=parcel, invoice_no=f'INV-{seq_year}-{seq:06d}', branch=parcel.branch,
                    total_weight=total_weight, weight_charge=weight_charge,
                    additional_charges_total=additional_charges_total,
                    customs_tax=customs, grand_total=grand_total, created_by=request.user)
                TrackingEvent.objects.create(
                    parcel=parcel, status=Parcel.Status.ORDER_CREATED,
                    location=parcel.branch.location, note='Parcel registered at branch',
                    created_by=request.user)
                log(request.user, f'Created shipment {parcel.tracking_id} / {invoice.invoice_no} '
                                  f'(NPR {grand_total}, customs NPR {customs}) at {parcel.branch.name}')
            messages.success(request, f'Shipment {parcel.tracking_id} created.')
            return redirect('invoice_print', invoice.id)
        except (KeyError, ValueError, InvalidOperation, RateCard.DoesNotExist) as e:
            messages.error(request, f'Could not save shipment: {e}')

    # Prefill data when reopening a draft
    draft_json = None
    if draft:
        draft_json = {
            'boxes': [{
                'weight': str(b.weight), 'rate': str(b.rate_per_kg),
                'items': [{'name': i.name, 'qty': i.quantity} for i in b.items.all()],
                'charges': [{'name': c.name, 'weight': str(c.weight) if c.weight is not None else '',
                             'price': str(c.price)} for c in b.charges.all()],
            } for b in draft.boxes.all()],
        }
    return render(request, 'panel/shipment_form.html', {
        'rates': rates, 'customs_tax': CUSTOMS_TAX,
        'draft': draft, 'draft_json': draft_json,
    })


@panel_required
def drafts(request):
    """All unfinished shipments for this branch (owner sees every branch)."""
    items = branch_scope(Parcel.objects.filter(is_draft=True).select_related('branch'),
                         request.user)
    return render(request, 'panel/drafts.html', {'drafts': items})


@panel_required
def draft_delete(request, pk):
    draft = get_object_or_404(
        branch_scope(Parcel.objects.filter(is_draft=True), request.user), pk=pk)
    if request.method == 'POST':
        tid = draft.tracking_id
        draft.delete()  # safe: drafts never have an invoice
        log(request.user, f'Draft deleted: {tid}')
        messages.success(request, f'Draft {tid} deleted.')
    return redirect('drafts')


@panel_required
def sender_lookup(request):
    """Auto-fill: if this phone has shipped before, return their details."""
    phone = request.GET.get('phone', '').strip()
    p = Parcel.objects.filter(sender_phone=phone).order_by('-created_at').first() if phone else None
    if not p:
        return JsonResponse({'found': False})
    return JsonResponse({'found': True, 'name': p.sender_name,
                         'address': p.sender_address, 'email': p.sender_email})


# ---------- parcels ----------

@panel_required
def parcel_list(request):
    q = request.GET.get('q', '').strip()
    parcels = branch_scope(Parcel.objects.filter(is_draft=False)
                           .select_related('branch', 'invoice'), request.user)
    if q:
        parcels = parcels.filter(
            Q(tracking_id__icontains=q) | Q(sender_phone__icontains=q) |
            Q(sender_name__icontains=q) | Q(receiver_name__icontains=q) |
            Q(invoice__invoice_no__icontains=q))
    return render(request, 'panel/parcel_list.html', {'parcels': parcels[:100], 'q': q})


@panel_required
def parcel_detail(request, pk):
    parcel = get_object_or_404(branch_scope(Parcel.objects.all(), request.user), pk=pk)
    if parcel.is_draft:
        return redirect('draft_edit', pk=parcel.pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'status':
            new = request.POST['status']
            old = parcel.status
            if new != old:
                parcel.status = new
                parcel.save()
                TrackingEvent.objects.create(
                    parcel=parcel, status=new,
                    location=request.POST.get('location') or parcel.branch.location,
                    note=request.POST.get('note', ''), created_by=request.user)
                log(request.user, f'{parcel.tracking_id}: status {old} → {new}')
                messages.success(request, 'Status updated.')
        elif action == 'carrier':
            parcel.carrier = request.POST['carrier']
            parcel.carrier_tracking_no = request.POST.get('carrier_tracking_no', '')
            parcel.save()
            log(request.user, f'{parcel.tracking_id}: handed to {parcel.get_carrier_display()} '
                              f'({parcel.carrier_tracking_no})')
            messages.success(request, 'Carrier handover saved.')
        elif action == 'customs':
            # Owner-only: adjust customs tax for a special customer.
            # Weight/items/rates stay locked — ONLY customs and the grand
            # total change, and the change is written to the audit log.
            if not request.user.is_owner:
                return HttpResponseForbidden('Only the owner can adjust customs tax.')
            try:
                new_customs = Decimal(request.POST['customs_tax'])
                if new_customs < 0:
                    raise InvalidOperation
            except (KeyError, InvalidOperation):
                messages.error(request, 'Enter a valid customs amount.')
                return redirect('parcel_detail', pk=pk)
            inv = parcel.invoice
            old_customs = inv.customs_tax
            inv.customs_tax = new_customs.quantize(Decimal('0.01'))
            inv.grand_total = inv.weight_charge + inv.additional_charges_total + inv.customs_tax
            inv.save()
            log(request.user, f'{parcel.tracking_id}: customs tax NPR {old_customs} → '
                              f'NPR {inv.customs_tax} (grand total NPR {inv.grand_total})')
            messages.success(request, f'Customs tax updated to NPR {inv.customs_tax}.')
        return redirect('parcel_detail', pk=pk)
    return render(request, 'panel/parcel_detail.html',
                  {'parcel': parcel, 'statuses': Parcel.Status.choices, 'carriers': Parcel.CARRIERS})


@panel_required
def parcel_edit(request, pk):
    """Edit ONLY names, addresses and contact details after a shipment is
    saved. Items, weights, rates and charges are permanently locked."""
    parcel = get_object_or_404(
        branch_scope(Parcel.objects.filter(is_draft=False), request.user), pk=pk)
    if request.method == 'POST':
        before = (f'{parcel.sender_name}/{parcel.sender_phone} → '
                  f'{parcel.receiver_name}/{parcel.receiver_phone}')
        parcel.sender_name = request.POST.get('sender_name', parcel.sender_name).strip()
        parcel.sender_phone = request.POST.get('sender_phone', parcel.sender_phone).strip()
        parcel.sender_address = request.POST.get('sender_address', parcel.sender_address).strip()
        parcel.sender_email = request.POST.get('sender_email', parcel.sender_email).strip()
        parcel.receiver_name = request.POST.get('receiver_name', parcel.receiver_name).strip()
        parcel.receiver_phone = request.POST.get('receiver_phone', parcel.receiver_phone).strip()
        parcel.receiver_address = request.POST.get('receiver_address', parcel.receiver_address).strip()
        parcel.save()
        log(request.user, f'{parcel.tracking_id}: contact details edited (was {before})')
        messages.success(request, 'Contact details updated. Items and pricing are unchanged (locked).')
        return redirect('parcel_detail', pk=pk)
    return render(request, 'panel/parcel_edit.html', {'parcel': parcel})


@panel_required
def invoice_print(request, pk):
    invoice = get_object_or_404(branch_scope(Invoice.objects.select_related('parcel', 'branch'),
                                             request.user), pk=pk)
    return render(request, 'panel/invoice_print.html',
                  {'inv': invoice, 'terms': TERMS_AND_CONDITIONS})


# ---------- HUB commercial invoice (Excel download) ----------
# ---------- HUB commercial invoice (Excel download) ----------
# Drop this in place of your existing hub_invoice_excel / hub_excel functions.
# It matches your urls.py entry:
#   path('panel/invoices/<int:pk>/hub-excel/', ops.hub_invoice_excel, name='hub_invoice_excel'),
# i.e. `pk` here is the INVOICE pk (not the parcel pk).

_ONES = ['', 'ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX', 'SEVEN', 'EIGHT', 'NINE',
         'TEN', 'ELEVEN', 'TWELVE', 'THIRTEEN', 'FOURTEEN', 'FIFTEEN', 'SIXTEEN',
         'SEVENTEEN', 'EIGHTEEN', 'NINETEEN']
_TENS = ['', '', 'TWENTY', 'THIRTY', 'FORTY', 'FIFTY', 'SIXTY', 'SEVENTY', 'EIGHTY', 'NINETY']


def _words_below_1000(n):
    parts = []
    if n >= 100:
        parts.append(_ONES[n // 100] + ' HUNDRED')
        n %= 100
    if n >= 20:
        parts.append(_TENS[n // 10] + (' ' + _ONES[n % 10] if n % 10 else ''))
    elif n:
        parts.append(_ONES[n])
    return ' '.join(parts)


def amount_in_words_usd(amount):
    """e.g. Decimal('150') -> 'ONE HUNDRED FIFTY ONLY'."""
    n = int(amount)
    if n == 0:
        return 'ZERO ONLY'
    parts = []
    for div, name in ((1_000_000, 'MILLION'), (1_000, 'THOUSAND')):
        if n >= div:
            parts.append(_words_below_1000(n // div) + f' {name}')
            n %= div
    if n:
        parts.append(_words_below_1000(n))
    return ' '.join(parts) + ' ONLY'


def _update_invoice(parcel, data, user):
    """Recalculate an existing invoice in place after a shipment edit.
    Same invoice number, same tracking ID — only the figures change.
    Payments are NOT touched here (there is no payment record on edits)."""
    inv = parcel.invoice
    old_total = inv.grand_total
    inv.total_weight = sum(b['weight'] for b in data['boxes'])
    inv.weight_charge = sum((b['weight'] * b['rate']).quantize(Decimal('0.01')) for b in data['boxes'])
    inv.additional_charges_total = sum(
        (c[2] for b in data['boxes'] for c in b['charges']), Decimal('0.00'))
    inv.customs_tax = data['customs_tax']
    inv.grand_total = inv.weight_charge + inv.additional_charges_total + inv.customs_tax
    inv.save()
    log(user, f'{parcel.tracking_id}: shipment edited — invoice {inv.invoice_no} '
              f'NPR {old_total} → NPR {inv.grand_total}')
    return inv


@panel_required
def hub_invoice_excel(request, pk):
    """Download the HUB commercial invoice as an editable Excel file.
    `pk` is the Invoice pk (matches urls.py: panel/invoices/<pk>/hub-excel/).
    Subtotals and totals are real formulas — edit qty/value after download
    and everything recalculates."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.properties import PageSetupProperties
    except ImportError:
        messages.error(request, 'openpyxl is not installed — run: pip install openpyxl')
        return redirect('hub_print', pk=pk)

    invoice = get_object_or_404(branch_scope(
        Invoice.objects.select_related('parcel', 'parcel__branch'), request.user), pk=pk)
    parcel = invoice.parcel
    boxes = list(parcel.boxes.prefetch_related('items').all())
    invoice_no = invoice.invoice_no

    wb = Workbook()
    ws = wb.active
    ws.title = 'HUB Invoice'

    # ---- styles ----
    yellow = PatternFill('solid', start_color='FFFF00')
    orange = PatternFill('solid', start_color='F4B183')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    widths = {'A': 5, 'B': 42, 'C': 9, 'D': 16, 'E': 12, 'F': 13, 'G': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    def cell(ref, value='', bold=False, fill=None, align=left, size=12):
        c = ws[ref]
        c.value = value
        c.font = Font(name='Times New Roman', size=size, bold=bold)
        if fill:
            c.fill = fill
        c.alignment = align
        return c

    def border_range(ref):
        for row in ws[ref]:
            for c in row:
                c.border = border

    # ---- FROM / INVOICE header ----
    ws.merge_cells('A1:C1'); cell('A1', 'FROM', bold=True, fill=yellow)
    ws.merge_cells('D1:G2'); cell('D1', 'INVOICE', bold=True, fill=yellow, align=center, size=24)
    ws.merge_cells('B2:C2'); cell('A2', 'Name:-', bold=True); cell('B2', parcel.sender_name)
    ws.merge_cells('B3:C3'); cell('A3', 'Company', bold=True); cell('B3', 'COURIER SPOT PVT. LTD.')
    cell('D3', 'INVOICE NO', bold=True, fill=yellow)
    ws.merge_cells('E3:G3'); cell('E3', invoice_no, bold=True)
    ws.merge_cells('B4:C4'); cell('A4', 'Address:-', bold=True); cell('B4', parcel.branch.address)
    cell('D4', 'DATE :', bold=True, fill=yellow)
    ws.merge_cells('E4:G4'); cell('E4', timezone.localtime(invoice.created_at).strftime('%d.%m.%Y'))
    ws.merge_cells('B5:C5'); cell('A5', 'City & Postal code', bold=True); cell('B5', parcel.branch.location)
    cell('D5', 'TOTAL  WEIGHT:', bold=True, fill=yellow)
    cell('E5', float(invoice.total_weight), align=center)
    cell('F5', 'SIZE', bold=True, align=center); cell('G5', '')
    ws.merge_cells('B6:C6'); cell('A6', 'Country:-', bold=True); cell('B6', 'NEPAL')
    cell('D6', 'HAWB', bold=True, fill=yellow)
    ws.merge_cells('E6:G6'); cell('E6', parcel.carrier_tracking_no or '')
    ws.merge_cells('B7:C7'); cell('A7', 'Tel:-', bold=True); cell('B7', parcel.branch.phone)
    cell('D7', 'TOTAL BOX:-', bold=True, fill=yellow)
    ws.merge_cells('E7:G7'); cell('E7', len(boxes))

    # ---- TO ----
    ws.merge_cells('A8:C8'); cell('A8', 'TO', bold=True, fill=yellow)
    cell('D8', 'CARRIER', bold=True, fill=yellow)
    ws.merge_cells('E8:G8')
    cell('E8', parcel.get_carrier_display().upper() if parcel.carrier else 'COURIER',
         bold=True, align=center)
    ws.merge_cells('B9:C9');  cell('A9', 'Name:-', bold=True);  cell('B9', parcel.receiver_name)
    ws.merge_cells('B10:C10'); cell('A10', 'Company', bold=True); cell('B10', parcel.receiver_name)
    ws.merge_cells('B11:C11'); cell('A11', 'Address:-', bold=True); cell('B11', parcel.receiver_address)
    ws.merge_cells('B12:C12'); cell('A12', 'Country:-', bold=True); cell('B12', parcel.destination_country.upper())
    ws.merge_cells('B13:C13'); cell('A13', 'Tel:-', bold=True); cell('B13', parcel.receiver_phone)
    border_range('A1:G13')

    # ---- goods table ----
    hdr = 14
    for col, text in zip('ABCDEFG', ['SN', 'Description of goods', 'KG', 'COUNTRY OF\nORIGIN',
                                     'PKT BOX\nPCS', 'UNIT VALUE\n(IN USD)', 'SUB\nTOTAL']):
        cell(f'{col}{hdr}', text, bold=True, fill=yellow, align=center)
    ws.row_dimensions[hdr].height = 34

    r = hdr + 1
    item_rows = []
    for bi, box in enumerate(boxes, 1):
        ws.merge_cells(f'A{r}:G{r}')
        cell(f'A{r}', f'BOX-{bi:02d}', bold=True, fill=orange, align=center)
        r += 1
        for si, item in enumerate(box.items.all(), 1):
            cell(f'A{r}', si, align=center)
            cell(f'B{r}', item.name, bold=True)
            if si == 1:
                cell(f'C{r}', float(box.weight), align=center)
            cell(f'D{r}', 'NEPAL', bold=True, align=center)
            cell(f'E{r}', item.quantity, bold=True, align=center)
            cell(f'F{r}', float(item.unit_value_usd), bold=True, align=center)
            cell(f'G{r}', f'=E{r}*F{r}', bold=True, align=center)  # live formula
            item_rows.append(r)
            r += 1

    # a few blank rows so more items can be added by hand after download
    for _ in range(4):
        for col in 'ABCDEFG':
            cell(f'{col}{r}', '')
        cell(f'G{r}', f'=IF(E{r}*F{r}=0,"",E{r}*F{r})', align=center)
        r += 1

    first, last = hdr + 1, r - 1
    cell(f'D{r}', 'TOTAL QTY', bold=True, fill=yellow, align=center)
    cell(f'E{r}', f'=SUM(E{first}:E{last})', bold=True, fill=yellow, align=center)
    cell(f'F{r}', 'TOTAL VALUE', bold=True, fill=yellow, align=center)
    cell(f'G{r}', f'=SUM(G{first}:G{last})', bold=True, fill=yellow, align=center)
    border_range(f'A{hdr}:G{r}')
    r += 1

    # ---- declarations ----
    total_value = sum((i.subtotal_usd for b in boxes for i in b.items.all()), Decimal('0.00'))
    ws.merge_cells(f'A{r}:G{r}')
    cell(f'A{r}', f'IN WORDS (USD): {amount_in_words_usd(total_value)}', bold=True); r += 1
    for line in ['REASON FOR EXPORT: I DECLARE THAT ABOVE GOODS ARE FOR GIFT PURPOSE AND PERSONAL USE',
                 'DECLARED GOODS VALUE IS ONLY FOR CUSTOMS PURPOSE',
                 'TERMS OF DELIVERY : PREPAID']:
        ws.merge_cells(f'A{r}:G{r}')
        cell(f'A{r}', line, bold=True)
        border_range(f'A{r}:G{r}')
        r += 1

    # ---- signature block ----
    r += 1
    cell(f'D{r}', 'Name', bold=True); ws.merge_cells(f'E{r}:G{r}')
    cell(f'E{r}', parcel.sender_name.upper(), bold=True); border_range(f'D{r}:G{r}'); r += 1
    cell(f'D{r}', 'Date', bold=True); ws.merge_cells(f'E{r}:G{r}')
    cell(f'E{r}', timezone.localdate().strftime('%d.%m.%Y')); border_range(f'D{r}:G{r}'); r += 2
    cell(f'D{r}', 'SIGNATURE:', bold=True); ws.merge_cells(f'E{r}:G{r}'); cell(f'E{r}', '')
    border_range(f'D{r}:G{r}')

    # ---- print setup: fit to one page wide when printed from Excel ----
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f'A1:G{r}'

    # ---- respond as download ----
    buf = BytesIO()
    wb.save(buf)
    safe_name = re.sub(r'[^\w\s-]', '', parcel.receiver_name).strip() or 'HUB'
    filename = f"{safe_name} {invoice_no}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    log(request.user, f'{parcel.tracking_id}: HUB Excel invoice downloaded')
    return resp
# ---------- expenses ----------

@panel_required
def expenses(request):
    if request.method == 'POST':
        Expense.objects.create(
            branch=request.user.branch or Branch.objects.first(),
            description=request.POST['description'], category=request.POST['category'],
            amount=Decimal(request.POST['amount']), date=request.POST['date'] or date.today(),
            bill_image=request.FILES.get('bill_image'), created_by=request.user)
        log(request.user, f"Expense added: {request.POST['description']} NPR {request.POST['amount']}")
        messages.success(request, 'Expense recorded.')
        return redirect('expenses')
    items = branch_scope(Expense.objects.select_related('branch'), request.user)[:100]
    total = branch_scope(Expense.objects.all(), request.user).aggregate(s=Sum('amount'))['s'] or 0
    return render(request, 'panel/expenses.html',
                  {'expenses': items, 'total': total, 'categories': Expense.CATEGORIES})


# ---------- owner-only: finance, branches, staff, rates ----------

@owner_required
def finance(request):
    """Income − expenses per branch. ONLY the owner ever sees this."""
    rows = []
    for b in Branch.objects.all():
        income = b.invoices.aggregate(s=Sum('grand_total'))['s'] or Decimal('0')
        spent = b.expenses.aggregate(s=Sum('amount'))['s'] or Decimal('0')
        rows.append({'branch': b, 'income': income, 'spent': spent, 'balance': income - spent})
    totals = {
        'income': sum(r['income'] for r in rows),
        'spent': sum(r['spent'] for r in rows),
        'balance': sum(r['balance'] for r in rows),
    }
    return render(request, 'panel/finance.html',
                  {'rows': rows, 'totals': totals, 'audit': AuditLog.objects.select_related('user')[:30]})


@owner_required
def branches(request):
    if request.method == 'POST':
        Branch.objects.create(
            name=request.POST['name'], location=request.POST['location'],
            address=request.POST['address'], phone=request.POST['phone'],
            email=request.POST.get('email', ''))
        log(request.user, f"Branch created: {request.POST['name']}")
        messages.success(request, 'Branch created.')
        return redirect('branches')
    return render(request, 'panel/branches.html',
                  {'branches': Branch.objects.all(), 'staff': User.objects.exclude(role='CUSTOMER')})


@owner_required
def staff_create(request):
    form = StaffCreateForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        u = form.save()
        log(request.user, f'Panel account created: {u.username} ({u.get_role_display()})')
        messages.success(request, f'Account for {u.username} created.')
        return redirect('branches')
    return render(request, 'panel/staff_form.html', {'form': form})


@owner_required
def rates_manage(request):
    if request.method == 'POST':
        RateCard.objects.update_or_create(
            country=request.POST['country'],
            defaults={'rate_per_kg': Decimal(request.POST['rate_per_kg']),
                      'delivery_days': request.POST['delivery_days'],
                      'carriers': request.POST['carriers'],
                      'flag': request.POST.get('flag', '')})
        log(request.user, f"Rate saved: {request.POST['country']} NPR {request.POST['rate_per_kg']}/kg")
        messages.success(request, 'Rate saved — public website updated instantly.')
        return redirect('rates_manage')
    return render(request, 'panel/rates.html',
                  {'rates': RateCard.objects.all(), 'restricted': RestrictedItem.objects.all()})


@owner_required
def restricted_add(request):
    if request.method == 'POST':
        RestrictedItem.objects.create(country=request.POST['country'],
                                      item=request.POST['item'],
                                      note=request.POST.get('note', ''))
        messages.success(request, 'Restricted item added.')
    return redirect('rates_manage')
